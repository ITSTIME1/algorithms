import pandas as pd
import openpyxl

# 사료급이
class 연구소_데이터_처리:
    def __init__(self) -> None:
        self.file_path = "/Users/itstime/algorithms/python-algo/연구소/"
        self.엑셀_기입_파일 = self.file_path + "2023_뱀장어_사육_수정.xlsx"
        self.뱀장어_사육_일지_파일이름 = self.file_path + "실뱀장어 사육일지(2023)_방류용.xlsx"
        
        self.뱀장어_마리수 = []
        self.뱀장어_폐사량 = []
        self.뱀장어_수온 = []
        self.뱀장어_DO = []
        self.뱀장어_PH = []
        self.뱀장어_NH3 = []
        self.뱀장어_NO2 = []
        self.뱀장어_NO3 = []
        self.뱀장어_비고 = []
        self.뱀장어_공통 = []
        self.뱀장어_평균길이 = []
        self.뱀장어_총중량 = []
        
    def test(self):
        try:
            
            df = pd.read_excel(self.뱀장어_사육_수정)
            df1 = pd.read_excel(self.뱀장어_사육_일지)
            print(df)
            print(df1)
            print("성공")
        except Exception as e:
            print(e)
        
        
    def 사료급이(self):
        # 사료 급이 오전/오후로 나누자.
        
        
        pass

    # 물
    def 물(self):
        pass


    # 수온
    def 수온(self):
        xls = pd.ExcelFile(self.뱀장어_사육_일지_파일이름)  # 파일 경로를 실제 파일 경로로 수정해주세요.

        # 시트 이름 가져오기
        시트들 = xls.sheet_names
        # print(시트들)
        # 시트들에서 입식수량만 가지오자
        for i in 시트들:   
            df = pd.read_excel(self.뱀장어_사육_일지_파일이름, header=None, sheet_name=i)
            수온 = tuple(df[11][5:7])
            self.뱀장어_수온.append(수온)
        
        index = 0;
        # 엑셀 파일 열기
        wb = openpyxl.load_workbook(self.file_path + "새로운기입파일.xlsx")  # 파일 경로를 적절히 수정하세요
        ws = wb.active
        # # 데이터 입력
        index = 0
        for i in range(3, 42):
            ws.cell(row = 72, column = i, value = self.뱀장어_수온[index][0])
            ws.cell(row = 73, column = i, value = self.뱀장어_수온[index][1])
            # ws.cell(row = 15, column = i, value = self.뱀장어_폐사량[index][0] + self.뱀장어_폐사량[index][1])
            index+=1
            

        # # 파일 저장
        wb.save(self.file_path + '새로운기입파일.xlsx') # 
        # print(self.뱀장어_마리수)

    # DO
    def DO(self):
        xls = pd.ExcelFile(self.뱀장어_사육_일지_파일이름)  # 파일 경로를 실제 파일 경로로 수정해주세요.

        # 시트 이름 가져오기
        시트들 = xls.sheet_names
        # print(시트들)
        # 시트들에서 입식수량만 가지오자
        for i in 시트들:   
            df = pd.read_excel(self.뱀장어_사육_일지_파일이름, header=None, sheet_name=i)
            DO = tuple(df[13][5:7])
            self.뱀장어_DO.append(DO)
        print(self.뱀장어_DO)
        index = 0;
        # 엑셀 파일 열기
        wb = openpyxl.load_workbook(self.file_path + "새로운기입파일.xlsx")  # 파일 경로를 적절히 수정하세요
        ws = wb.active
        # # 데이터 입력
        index = 0
        for i in range(3, 42):
            ws.cell(row = 75, column = i, value = self.뱀장어_DO[index][0])
            ws.cell(row = 76, column = i, value = self.뱀장어_DO[index][1])
            # ws.cell(row = 15, column = i, value = self.뱀장어_폐사량[index][0] + self.뱀장어_폐사량[index][1])
            index+=1
            

        # # 파일 저장
        wb.save(self.file_path + '새로운기입파일.xlsx') # 
        # print(self.뱀장어_마리수)
    
    # PH 
    def PH(self):
        xls = pd.ExcelFile(self.뱀장어_사육_일지_파일이름)  # 파일 경로를 실제 파일 경로로 수정해주세요.

        # 시트 이름 가져오기
        시트들 = xls.sheet_names
        # print(시트들)
        # 시트들에서 입식수량만 가지오자
        for i in 시트들:   
            df = pd.read_excel(self.뱀장어_사육_일지_파일이름, header=None, sheet_name=i)
            PH = tuple(df[12][5:7])
            self.뱀장어_PH.append(PH)
        index = 0;
        # 엑셀 파일 열기
        wb = openpyxl.load_workbook(self.file_path + "새로운기입파일.xlsx")  # 파일 경로를 적절히 수정하세요
        ws = wb.active
        # # 데이터 입력
        index = 0
        for i in range(3, 42):
            ws.cell(row = 78, column = i, value = self.뱀장어_PH[index][0])
            ws.cell(row = 79, column = i, value = self.뱀장어_PH[index][1])
            # ws.cell(row = 15, column = i, value = self.뱀장어_폐사량[index][0] + self.뱀장어_폐사량[index][1])
            index+=1
            

        # # 파일 저장
        wb.save(self.file_path + '새로운기입파일.xlsx') # 
        # print(self.뱀장어_마리수)

    # NH3
    def NH3(self):
        xls = pd.ExcelFile(self.뱀장어_사육_일지_파일이름)  # 파일 경로를 실제 파일 경로로 수정해주세요.

        # 시트 이름 가져오기
        시트들 = xls.sheet_names
        # print(시트들)
        # 시트들에서 입식수량만 가지오자
        for i in 시트들:   
            df = pd.read_excel(self.뱀장어_사육_일지_파일이름, header=None, sheet_name=i)
            NH3 = tuple(df[14][5:7])
            self.뱀장어_NH3.append(NH3)
        index = 0;
        # 엑셀 파일 열기
        wb = openpyxl.load_workbook(self.file_path + "새로운기입파일.xlsx")  # 파일 경로를 적절히 수정하세요
        ws = wb.active
        # # 데이터 입력
        index = 0
        for i in range(3, 42):
            ws.cell(row = 81, column = i, value = self.뱀장어_NH3[index][0])
            ws.cell(row = 82, column = i, value = self.뱀장어_NH3[index][1])
            # ws.cell(row = 15, column = i, value = self.뱀장어_폐사량[index][0] + self.뱀장어_폐사량[index][1])
            index+=1
            

        # # 파일 저장
        wb.save(self.file_path + '새로운기입파일.xlsx') # 
        # print(self.뱀장어_마리수)

    # NO2
    def NO2(self):
        xls = pd.ExcelFile(self.뱀장어_사육_일지_파일이름)  # 파일 경로를 실제 파일 경로로 수정해주세요.

        # 시트 이름 가져오기
        시트들 = xls.sheet_names
        # print(시트들)
        # 시트들에서 입식수량만 가지오자
        for i in 시트들:   
            df = pd.read_excel(self.뱀장어_사육_일지_파일이름, header=None, sheet_name=i)
            NO2 = tuple(df[15][5:7])
            self.뱀장어_NO2.append(NO2)
        index = 0;
        # 엑셀 파일 열기
        wb = openpyxl.load_workbook(self.file_path + "새로운기입파일.xlsx")  # 파일 경로를 적절히 수정하세요
        ws = wb.active
        # # 데이터 입력
        index = 0
        for i in range(3, 42):
            ws.cell(row = 84, column = i, value = self.뱀장어_NO2[index][0])
            ws.cell(row = 85, column = i, value = self.뱀장어_NO2[index][1])
            # ws.cell(row = 15, column = i, value = self.뱀장어_폐사량[index][0] + self.뱀장어_폐사량[index][1])
            index+=1
            

        # # 파일 저장
        wb.save(self.file_path + '새로운기입파일.xlsx') # 
        # print(self.뱀장어_마리수)

    # NO3
    def NO3(self):
        xls = pd.ExcelFile(self.뱀장어_사육_일지_파일이름)  # 파일 경로를 실제 파일 경로로 수정해주세요.

        # 시트 이름 가져오기
        시트들 = xls.sheet_names
        # print(시트들)
        # 시트들에서 입식수량만 가지오자
        for i in 시트들:   
            df = pd.read_excel(self.뱀장어_사육_일지_파일이름, header=None, sheet_name=i)
            NO3 = tuple(df[16][5:7])
            self.뱀장어_NO3.append(NO3)
        index = 0;
        # 엑셀 파일 열기
        wb = openpyxl.load_workbook(self.file_path + "새로운기입파일.xlsx")  # 파일 경로를 적절히 수정하세요
        ws = wb.active
        # # 데이터 입력
        index = 0
        for i in range(3, 42):
            ws.cell(row = 87, column = i, value = self.뱀장어_NO3[index][0])
            ws.cell(row = 88, column = i, value = self.뱀장어_NO3[index][1])
            # ws.cell(row = 15, column = i, value = self.뱀장어_폐사량[index][0] + self.뱀장어_폐사량[index][1])
            index+=1
            

        # # 파일 저장
        wb.save(self.file_path + '새로운기입파일.xlsx') # 
        # print(self.뱀장어_마리수)

    # 비고
    def 비고_약욕_및_기타(self):
        xls = pd.ExcelFile(self.뱀장어_사육_일지_파일이름)  # 파일 경로를 실제 파일 경로로 수정해주세요.

        # 시트 이름 가져오기
        시트들 = xls.sheet_names
        # print(시트들)
        # 시트들에서 입식수량만 가지오자
        for i in 시트들:   
            df = pd.read_excel(self.뱀장어_사육_일지_파일이름, header=None, sheet_name=i)
            비고 = tuple(df[17][5:7])
            self.뱀장어_비고.append(비고)
        index = 0;
        # 엑셀 파일 열기
        wb = openpyxl.load_workbook(self.file_path + "새로운기입파일.xlsx")  # 파일 경로를 적절히 수정하세요
        ws = wb.active
        # # 데이터 입력
        index = 0
        for i in range(3, 42):
            ws.cell(row = 92, column = i, value = self.뱀장어_비고[index][0])
            ws.cell(row = 93, column = i, value = self.뱀장어_비고[index][1])
            # ws.cell(row = 15, column = i, value = self.뱀장어_폐사량[index][0] + self.뱀장어_폐사량[index][1])
            index+=1
            

        # # 파일 저장
        wb.save(self.file_path + '새로운기입파일.xlsx') # 
        # print(self.뱀장어_마리수)

    # 공통데이터
    def 공통_데이터_추출(self):
        xls = pd.ExcelFile(self.뱀장어_사육_일지_파일이름)  # 파일 경로를 실제 파일 경로로 수정해주세요.

        # 시트 이름 가져오기
        시트들 = xls.sheet_names
        # print(시트들)
        # 시트들에서 입식수량만 가지오자
        for i in 시트들:   
            df = pd.read_excel(self.뱀장어_사육_일지_파일이름, header=None, sheet_name=i)
            공통 = "".join(list(df.iloc[11][1])[2:])
            self.뱀장어_공통.append(공통)
        # print(self.뱀장어_공통)
        index = 0;
        # 엑셀 파일 열기
        wb = openpyxl.load_workbook(self.file_path + "새로운기입파일.xlsx")  # 파일 경로를 적절히 수정하세요
        ws = wb.active
        # # 데이터 입력
        index = 0
        for i in range(3, 42):
            ws.cell(row = 94, column = i, value = self.뱀장어_공통[index])
            # ws.cell(row = 93, column = i, value = self.뱀장어_비고[index][1])
            # ws.cell(row = 15, column = i, value = self.뱀장어_폐사량[index][0] + self.뱀장어_폐사량[index][1])
            index+=1
            

        # # 파일 저장
        wb.save(self.file_path + '새로운기입파일.xlsx') # 

    # 폐사량
    def 폐사량(self):
        xls = pd.ExcelFile(self.뱀장어_사육_일지_파일이름)  # 파일 경로를 실제 파일 경로로 수정해주세요.

        # 시트 이름 가져오기
        시트들 = xls.sheet_names
        # print(시트들)
        # 시트들에서 입식수량만 가지오자
        for i in 시트들:   
            df = pd.read_excel(self.뱀장어_사육_일지_파일이름, header=None, sheet_name=i)
            폐사 = tuple(df[4][5:7])
            self.뱀장어_폐사량.append(폐사)
        
        index = 0;
        # 엑셀 파일 열기
        wb = openpyxl.load_workbook(self.file_path + "새로운기입파일.xlsx")  # 파일 경로를 적절히 수정하세요
        ws = wb.active
        # # 데이터 입력
        index = 0
        for i in range(3, 42):
            ws.cell(row = 13, column = i, value = self.뱀장어_폐사량[index][0])
            ws.cell(row = 14, column = i, value = self.뱀장어_폐사량[index][1])
            ws.cell(row = 15, column = i, value = self.뱀장어_폐사량[index][0] + self.뱀장어_폐사량[index][1])
            index+=1
            

        # # 파일 저장
        wb.save(self.file_path + '새로운기입파일.xlsx') # 
        # print(self.뱀장어_마리수)
    
    # 마리수
    def 마리수(self):
        # 마리수가 가장 간단하니까 마리수 부터 하자.
        # Excel 파일 열기
        xls = pd.ExcelFile(self.뱀장어_사육_일지_파일이름)  # 파일 경로를 실제 파일 경로로 수정해주세요.

        # 시트 이름 가져오기
        시트들 = xls.sheet_names
        # print(시트들)
        # 시트들에서 입식수량만 가지오자
        for i in 시트들:   
            df = pd.read_excel(self.뱀장어_사육_일지_파일이름, header=None, sheet_name=i)
            입식수량마리 = tuple(df[3][5:7])
            self.뱀장어_마리수.append(입식수량마리)
    
        # 엑셀 파일 열기
        wb = openpyxl.load_workbook(self.엑셀_기입_파일)  # 파일 경로를 적절히 수정하세요
        ws = wb.active
        # 데이터 입력
        index = 0
        for i in range(3, 42):
            ws.cell(row = 10, column = i, value = self.뱀장어_마리수[index][0])
            ws.cell(row = 11, column = i, value = self.뱀장어_마리수[index][1])
            ws.cell(row = 12, column = i, value = self.뱀장어_마리수[index][0] + self.뱀장어_마리수[index][1])
            index+=1
            

        # 파일 저장
        wb.save(self.file_path + '새로운기입파일.xlsx')
        print(self.뱀장어_마리수)
            
        # for i in df_copy.columns[2:41]:
        #     print(df_copy[i][9])
        #     print(df_copy[i][10])
        #     print("next")
        
        # print(df_copy[9:11])
        # df.to_excel(self.file_path + '새로운기입파일.xlsx', index=False)  # 파일 경로와 이름을 적절히 수정하세요
            
            
    # 평균길이
    def 평균길이(self):
        # 마리수가 가장 간단하니까 마리수 부터 하자.
        # Excel 파일 열기
        xls = pd.ExcelFile(self.뱀장어_사육_일지_파일이름)  # 파일 경로를 실제 파일 경로로 수정해주세요.

        # 시트 이름 가져오기
        시트들 = xls.sheet_names
        # print(시트들)
        # 시트들에서 입식수량만 가지오자
        for i in 시트들:   
            df = pd.read_excel(self.뱀장어_사육_일지_파일이름, header=None, sheet_name=i)
            평균길이 = tuple(df[7][5:7])
            self.뱀장어_평균길이.append(평균길이)
        
        # # 엑셀 파일 열기
        wb = openpyxl.load_workbook(self.file_path + "새로운기입파일.xlsx")  # 파일 경로를 적절히 수정하세요
        ws = wb.active
        # 데이터 입력
        index = 0
        for i in range(3, 42):
            ws.cell(row = 4, column = i, value = self.뱀장어_평균길이[index][0])
            index+=1
            

        # # 파일 저장
        wb.save(self.file_path + '새로운기입파일.xlsx')
        
    def 총중량(self):
        # 마리수가 가장 간단하니까 마리수 부터 하자.
        # Excel 파일 열기
        xls = pd.ExcelFile(self.뱀장어_사육_일지_파일이름)  # 파일 경로를 실제 파일 경로로 수정해주세요.

        # 시트 이름 가져오기
        시트들 = xls.sheet_names
        # print(시트들)
        # 시트들에서 입식수량만 가지오자
        for i in 시트들:   
            df = pd.read_excel(self.뱀장어_사육_일지_파일이름, header=None, sheet_name=i)
            총중량 = tuple(df[8][5:7])
            self.뱀장어_총중량.append(총중량)
        
        # # 엑셀 파일 열기
        wb = openpyxl.load_workbook(self.file_path + "새로운기입파일.xlsx")  # 파일 경로를 적절히 수정하세요
        ws = wb.active
        # 데이터 입력
        index = 0
        for i in range(3, 42):
            ws.cell(row = 6, column = i, value = self.뱀장어_총중량[index][0])
            ws.cell(row = 7, column = i, value = self.뱀장어_총중량[index][1])
            index+=1
            

        # # 파일 저장
        wb.save(self.file_path + '새로운기입파일.xlsx')


    
        


if __name__ == '__main__':
    연구소 = 연구소_데이터_처리()
    # 연구소.test()
    # 연구소.마리수() # 마리수완료, sum으로 바꾸기만 하면됨
    # 연구소.폐사량() # 폐사량 완료, 합계는 어떤걸로 할 지모름
    # 연구소.수온() # 수온완료
    # 연구소.DO()
    # 연구소.PH()
    # 연구소.NH3()
    # 연구소.NO2()
    # 연구소.NO3()
    # 연구소.비고_약욕_및_기타()
    # 연구소.공통_데이터_추출()
    # 연구소.평균길이()
    # 연구소.총중량()
    